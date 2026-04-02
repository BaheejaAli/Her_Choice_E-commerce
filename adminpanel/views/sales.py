from django.shortcuts import render
from django.db.models import Count, Sum
from orders.models import Order
from django.utils import timezone
from datetime import timedelta
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.core.paginator import Paginator
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.cache import never_cache
from django.contrib import messages


def is_admin(user):
    return user.is_staff or user.is_superuser


def get_filtered_orders(request):
    date_range = request.GET.get('date_range','month')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    orders = Order.objects.exclude(status__in=['cancelled','failed']).order_by("-created_at")

    today = timezone.now().date()
    if date_range == 'today':
        start_date = end_date = today
        orders = orders.filter(created_at__date=today)

    elif date_range == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
        orders = orders.filter(created_at__date__gte=start_date)

    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
        orders = orders.filter(created_at__date__gte=start_date)

    elif date_range == 'year':
        start_date = datetime(today.year, 1, 1).date()
        end_date = today
        orders = orders.filter(created_at__year=today.year)

    elif date_range == 'custom' and start_date and end_date:
        try:
            start_date_obj = datetime.strptime(start_date, "%d-%m-%Y").date()
            end_date_obj = datetime.strptime(end_date, "%d-%m-%Y").date()
            
            if start_date_obj > end_date_obj:
                messages.error(request, "Start date cannot be after end date.")
                # Fallback to current month if validation fails
                start_date = today.replace(day=1)
                end_date = today
                orders = orders.filter(created_at__date__gte=start_date)
            elif start_date_obj > today or end_date_obj > today:
                messages.error(request, "Dates cannot be in the future.")
                start_date = today.replace(day=1)
                end_date = today
                orders = orders.filter(created_at__date__gte=start_date)
            else:
                start_date, end_date = start_date_obj, end_date_obj
                orders = orders.filter(created_at__date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, "Invalid date format.")
            start_date = today.replace(day=1)
            end_date = today
            orders = orders.filter(created_at__date__gte=start_date)

    return orders, start_date, end_date

@login_required
@user_passes_test(is_admin)
@never_cache
def sales_report(request):
    orders, start_date, end_date = get_filtered_orders(request)
    total = orders.aggregate(
        sales_count=Count('id'),
        total_order_amount=Sum('total'),
        total_item_discount=Sum('discount'),
        total_coupon_discount=Sum('coupon_discount'),
    )

    total['total_item_discount'] = total['total_item_discount'] or 0
    total['total_coupon_discount'] = total['total_coupon_discount'] or 0

    total['total_discount'] = (
        total['total_item_discount'] + total['total_coupon_discount']
    )
    total['total_order_amount'] = total['total_order_amount'] or 0


    paginator = Paginator(orders, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj, 
        'is_paginated': page_obj.has_other_pages(),
        'page_range': paginator.get_elided_page_range(page_obj.number),
        'total':total,
        'start_date':start_date,
        'end_date':end_date
        }
    return render(request,"admin_panel/sales_report.html",context)

@login_required
@user_passes_test(is_admin)
@never_cache
def export_pdf(request):
    orders, start_date, end_date = get_filtered_orders(request)

    total = orders.aggregate(
        sales_count=Count('id'),
        total_order_amount=Sum('total'),
        total_item_discount=Sum('discount'),
        total_coupon_discount=Sum('coupon_discount')
    )
    
    total['total_item_discount'] = total['total_item_discount'] or 0
    total['total_coupon_discount'] = total['total_coupon_discount'] or 0

    total['total_discount'] = (
        total['total_item_discount'] + total['total_coupon_discount']
    )
    total['total_order_amount'] = total['total_order_amount'] or 0

    html_string = render_to_string("admin_panel/sales_report_pdf.html", {
        "orders": orders,
        "start_date": start_date,
        "end_date": end_date,
        "total": total,
        "now": timezone.now()
    })

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=sales_report.pdf"

    HTML(string=html_string).write_pdf(response)

    return response

@login_required
@user_passes_test(is_admin)
@never_cache
def export_excel(request):
    orders, start_date, end_date = get_filtered_orders(request)
    total = orders.aggregate(
        sales_count=Count('id'),
        total_order_amount=Sum('total'),
        total_item_discount=Sum('discount'),
        total_coupon_discount=Sum('coupon_discount')
    )
    total_orders = total['sales_count']
    total_amount = total['total_order_amount'] or 0
    total_coupon_discount = total['total_coupon_discount']

    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sales Report"

    headers = ["Order ID", "Date", "Customer", "Total", "Coupon Discount", "Tax + Delivery charge", "Final Amount", "Status"]

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Add data rows
    for row_num, order in enumerate(orders, 2):
        sheet.cell(row=row_num, column=1).value = order.orderid
        sheet.cell(row=row_num, column=2).value = order.created_at.strftime("%d-%m-%Y")
        sheet.cell(row=row_num, column=3).value = order.user.get_full_name if order.user else "Guest"
        sheet.cell(row=row_num, column=4).value = float(order.subtotal)
        sheet.cell(row=row_num, column=5).value = float(order.calculated_coupon_discount or 0)
        sheet.cell(row=row_num, column=6).value = order.tax + order.delivery_charge or 0
        sheet.cell(row=row_num, column=7).value = order.total
        sheet.cell(row=row_num, column=8).value = order.status

    summary_start_row = row_num + 2  # Leave one empty row

    sheet.cell(row=summary_start_row, column=1).value = "TOTAL ORDERS"
    sheet.cell(row=summary_start_row, column=2).value = total_orders

    sheet.cell(row=summary_start_row + 1, column=1).value = "TOTAL AMOUNT"
    sheet.cell(row=summary_start_row + 1, column=2).value = float(total_amount)

    sheet.cell(row=summary_start_row + 2, column=1).value = "TOTAL COUPON DISCOUNT"
    sheet.cell(row=summary_start_row + 2, column=2).value = float(total_coupon_discount)

    # Make summary bold
    for i in range(3):
        sheet.cell(row=summary_start_row + i, column=1).font = Font(bold=True)
    

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=sales_report.xlsx"

    workbook.save(response)
    return response



